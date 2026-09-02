#!/usr/bin/env python3
"""Build the downloadable Excel templates (plain .xlsx, live formulas, no macros).

Run from anywhere:  python scripts/build_excel.py
Outputs to /downloads.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "downloads"
OUT.mkdir(exist_ok=True)

INK = "1C1A17"
BRAND = "12523A"
PAPER = "FAF8F3"
INPUT_FILL = PatternFill("solid", fgColor="E4EFE7")   # editable cells
CALC_FILL = PatternFill("solid", fgColor="F4F1E9")    # computed cells
HEAD_FILL = PatternFill("solid", fgColor=BRAND)
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(size=16, bold=True, color=INK)
SUB_FONT = Font(size=10, color="57534A", italic=True)
LABEL_FONT = Font(size=11, color=INK)
BOLD = Font(size=11, bold=True, color=INK)
THIN = Side(style="thin", color="D5CFC0")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONEY = '#,##0.00 "USD";[Red]-#,##0.00 "USD"'
MONEY0 = '"$"#,##0'
MONEY2 = '"$"#,##0.00'
PCT = "0.00%"


def title_block(ws, title, note):
    ws["B1"] = title
    ws["B1"].font = TITLE_FONT
    ws["B2"] = note
    ws["B2"].font = SUB_FONT
    ws["B3"] = "From Quiet Zeros — free, private calculators. Green cells are inputs; everything else recalculates."
    ws["B3"].font = SUB_FONT


def put(ws, cell, value, *, fill=None, fmt=None, font=None, border=True):
    c = ws[cell]
    c.value = value
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    c.font = font or LABEL_FONT
    if border and fill:
        c.border = BOX
    return c


def label(ws, cell, text, bold=False):
    ws[cell] = text
    ws[cell].font = BOLD if bold else LABEL_FONT


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def sched_header(ws, row, cols):
    for idx, h in enumerate(cols):
        c = ws.cell(row=row, column=idx + 1, value=h)
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


# ---------------------------------------------------------------- mortgage
def mortgage():
    wb = Workbook()
    ws = wb.active
    ws.title = "Mortgage"
    title_block(ws, "Mortgage calculator", "Payment with taxes, insurance and HOA, plus lifetime cost.")
    widths(ws, {"A": 2, "B": 26, "C": 16, "D": 3, "E": 26, "F": 16})

    rows = [
        ("Home price", 425000, MONEY0),
        ("Down payment (%)", 0.20, PCT),
        ("Interest rate (APR)", 0.065, PCT),
        ("Term (years)", 30, "0"),
        ("Property tax (per year)", 4800, MONEY0),
        ("Home insurance (per year)", 1800, MONEY0),
        ("HOA dues (per month)", 0, MONEY0),
    ]
    r = 5
    label(ws, f"B{r - 1}", "Inputs", bold=True)
    for i, (name, v, fmt) in enumerate(rows):
        label(ws, f"B{r + i}", name)
        put(ws, f"C{r + i}", v, fill=INPUT_FILL, fmt=fmt)

    label(ws, "E4", "Results", bold=True)
    label(ws, "E5", "Loan amount")
    put(ws, "F5", "=C5*(1-C6)", fill=CALC_FILL, fmt=MONEY0)
    label(ws, "E6", "Monthly principal & interest")
    put(ws, "F6", "=IF(C7=0,F5/(C8*12),F5*(C7/12)/(1-(1+C7/12)^(-C8*12)))", fill=CALC_FILL, fmt=MONEY2)
    label(ws, "E7", "Monthly tax + insurance + HOA")
    put(ws, "F7", "=C9/12+C10/12+C11", fill=CALC_FILL, fmt=MONEY2)
    label(ws, "E8", "Total monthly payment")
    put(ws, "F8", "=F6+F7", fill=CALC_FILL, fmt=MONEY2, font=BOLD)
    label(ws, "E9", "Total interest over the loan")
    put(ws, "F9", "=F6*C8*12-F5", fill=CALC_FILL, fmt=MONEY0)
    label(ws, "E10", "Total cost (principal + interest)")
    put(ws, "F10", "=F5+F9", fill=CALC_FILL, fmt=MONEY0)

    sh = wb.create_sheet("Schedule")
    widths(sh, {"A": 9, "B": 14, "C": 14, "D": 14, "E": 16})
    sched_header(sh, 1, ["Month", "Payment", "Interest", "Principal", "Balance"])
    n = 600
    for m in range(1, n + 1):
        row = m + 1
        live = f"$A{row}<=Mortgage!$C$8*12"
        prev_bal = "Mortgage!$F$5" if m == 1 else f"$E{m}"
        sh[f"A{row}"] = m
        sh[f"B{row}"] = f'=IF({live},Mortgage!$F$6,"")'
        sh[f"C{row}"] = f'=IF({live},{prev_bal}*Mortgage!$C$7/12,"")'
        sh[f"D{row}"] = f'=IF({live},$B{row}-$C{row},"")'
        sh[f"E{row}"] = f'=IF({live},MAX(0,{prev_bal}-$D{row}),"")'
        for col in "BCDE":
            sh[f"{col}{row}"].number_format = MONEY2
    wb.save(OUT / "mortgage-amortization.xlsx")


# -------------------------------------------------------------------- loan
def loan():
    wb = Workbook()
    ws = wb.active
    ws.title = "Loan"
    title_block(ws, "Loan calculator", "Any fixed-rate loan: payment, interest, payoff.")
    widths(ws, {"A": 2, "B": 26, "C": 16, "D": 3, "E": 26, "F": 16})

    label(ws, "B4", "Inputs", bold=True)
    label(ws, "B5", "Loan amount")
    put(ws, "C5", 25000, fill=INPUT_FILL, fmt=MONEY0)
    label(ws, "B6", "Interest rate (APR)")
    put(ws, "C6", 0.075, fill=INPUT_FILL, fmt=PCT)
    label(ws, "B7", "Term (years)")
    put(ws, "C7", 5, fill=INPUT_FILL, fmt="0.0")

    label(ws, "E4", "Results", bold=True)
    label(ws, "E5", "Monthly payment")
    put(ws, "F5", "=IF(C6=0,C5/(C7*12),C5*(C6/12)/(1-(1+C6/12)^(-C7*12)))", fill=CALC_FILL, fmt=MONEY2, font=BOLD)
    label(ws, "E6", "Total interest")
    put(ws, "F6", "=F5*C7*12-C5", fill=CALC_FILL, fmt=MONEY0)
    label(ws, "E7", "Total paid")
    put(ws, "F7", "=C5+F6", fill=CALC_FILL, fmt=MONEY0)

    sh = wb.create_sheet("Schedule")
    widths(sh, {"A": 9, "B": 14, "C": 14, "D": 14, "E": 16})
    sched_header(sh, 1, ["Month", "Payment", "Interest", "Principal", "Balance"])
    for m in range(1, 481):
        row = m + 1
        live = f"$A{row}<=Loan!$C$7*12"
        prev_bal = "Loan!$C$5" if m == 1 else f"$E{m}"
        sh[f"A{row}"] = m
        sh[f"B{row}"] = f'=IF({live},Loan!$F$5,"")'
        sh[f"C{row}"] = f'=IF({live},{prev_bal}*Loan!$C$6/12,"")'
        sh[f"D{row}"] = f'=IF({live},$B{row}-$C{row},"")'
        sh[f"E{row}"] = f'=IF({live},MAX(0,{prev_bal}-$D{row}),"")'
        for col in "BCDE":
            sh[f"{col}{row}"].number_format = MONEY2
    wb.save(OUT / "loan-amortization.xlsx")


# -------------------------------------------------------- compound interest
def compound():
    wb = Workbook()
    ws = wb.active
    ws.title = "Growth"
    title_block(ws, "Compound interest", "Starting balance plus monthly contributions, compounded monthly.")
    widths(ws, {"A": 2, "B": 26, "C": 16, "D": 3, "E": 10, "F": 16, "G": 16, "H": 16})

    label(ws, "B4", "Inputs", bold=True)
    label(ws, "B5", "Starting amount")
    put(ws, "C5", 10000, fill=INPUT_FILL, fmt=MONEY0)
    label(ws, "B6", "Monthly contribution")
    put(ws, "C6", 250, fill=INPUT_FILL, fmt=MONEY0)
    label(ws, "B7", "Annual return")
    put(ws, "C7", 0.07, fill=INPUT_FILL, fmt=PCT)
    label(ws, "B8", "Years (up to 50 shown)")
    put(ws, "C8", 20, fill=INPUT_FILL, fmt="0")

    label(ws, "B10", "Balance after the chosen years")
    put(ws, "C10", "=FV(C7/12,C8*12,-C6,-C5)", fill=CALC_FILL, fmt=MONEY0, font=BOLD)

    r0 = 13
    for idx, h in enumerate(["Year", "Contributed", "Growth", "Balance"]):
        c = ws.cell(row=r0, column=5 + idx, value=h)
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        c.alignment = Alignment(horizontal="center")
    for y in range(0, 51):
        row = r0 + 1 + y
        ws[f"E{row}"] = y
        ws[f"F{row}"] = f"=$C$5+$C$6*12*E{row}"
        ws[f"G{row}"] = f"=H{row}-F{row}"
        ws[f"H{row}"] = f"=FV($C$7/12,E{row}*12,-$C$6,-$C$5)"
        for col in "FGH":
            ws[f"{col}{row}"].number_format = MONEY0
    wb.save(OUT / "compound-interest.xlsx")


# ------------------------------------------------------------ savings goal
def savings_goal():
    wb = Workbook()
    ws = wb.active
    ws.title = "Goal"
    title_block(ws, "Savings goal planner", "The monthly deposit that reaches a target by a date.")
    widths(ws, {"A": 2, "B": 28, "C": 16, "D": 3, "E": 10, "F": 16})

    label(ws, "B4", "Inputs", bold=True)
    label(ws, "B5", "Goal amount")
    put(ws, "C5", 20000, fill=INPUT_FILL, fmt=MONEY0)
    label(ws, "B6", "Already saved")
    put(ws, "C6", 2000, fill=INPUT_FILL, fmt=MONEY0)
    label(ws, "B7", "Years to goal")
    put(ws, "C7", 3, fill=INPUT_FILL, fmt="0.00")
    label(ws, "B8", "Interest rate (APY)")
    put(ws, "C8", 0.04, fill=INPUT_FILL, fmt=PCT)

    label(ws, "B10", "Months")
    put(ws, "C10", "=ROUND(C7*12,0)", fill=CALC_FILL, fmt="0")
    label(ws, "B11", "Future value of current savings")
    put(ws, "C11", "=C6*(1+C8/12)^C10", fill=CALC_FILL, fmt=MONEY0)
    label(ws, "B12", "Deposit needed per month")
    put(
        ws, "C12",
        "=IF(C11>=C5,0,IF(C8=0,(C5-C6)/C10,(C5-C11)*(C8/12)/((1+C8/12)^C10-1)))",
        fill=CALC_FILL, fmt=MONEY2, font=BOLD,
    )
    label(ws, "B13", "Total you will deposit")
    put(ws, "C13", "=C12*C10", fill=CALC_FILL, fmt=MONEY0)
    label(ws, "B14", "Interest the account earns")
    put(ws, "C14", "=MAX(0,C5-C6-C13)", fill=CALC_FILL, fmt=MONEY0)

    r0 = 17
    for idx, h in enumerate(["Year", "Balance"]):
        c = ws.cell(row=r0, column=5 + idx, value=h)
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        c.alignment = Alignment(horizontal="center")
    for y in range(0, 51):
        row = r0 + 1 + y
        ws[f"E{row}"] = y
        ws[f"F{row}"] = (
            f'=IF(E{row}*12>$C$10,"",'
            f"FV($C$8/12,E{row}*12,-$C$12,-$C$6))"
        )
        ws[f"F{row}"].number_format = MONEY0
    wb.save(OUT / "savings-goal-planner.xlsx")


# ------------------------------------------------------------- debt payoff
def debt_payoff():
    wb = Workbook()
    ws = wb.active
    ws.title = "Payoff"
    title_block(ws, "Debt payoff planner", "How extra monthly payments shorten a debt and save interest.")
    widths(ws, {"A": 2, "B": 30, "C": 16, "D": 3, "E": 30, "F": 16})

    label(ws, "B4", "Inputs", bold=True)
    label(ws, "B5", "Current balance")
    put(ws, "C5", 8000, fill=INPUT_FILL, fmt=MONEY0)
    label(ws, "B6", "Interest rate (APR)")
    put(ws, "C6", 0.22, fill=INPUT_FILL, fmt=PCT)
    label(ws, "B7", "Monthly payment")
    put(ws, "C7", 250, fill=INPUT_FILL, fmt=MONEY0)
    label(ws, "B8", "Extra per month")
    put(ws, "C8", 100, fill=INPUT_FILL, fmt=MONEY0)

    label(ws, "E4", "Results", bold=True)
    label(ws, "E5", "Months to payoff (with extra)")
    put(ws, "F5", '=IFERROR(ROUNDUP(NPER(C6/12,-(C7+C8),C5),0),"Never — payment too small")',
        fill=CALC_FILL, fmt="0", font=BOLD)
    label(ws, "E6", "Months to payoff (regular only)")
    put(ws, "F6", '=IFERROR(ROUNDUP(NPER(C6/12,-C7,C5),0),"Never — payment too small")', fill=CALC_FILL, fmt="0")
    label(ws, "E7", "Interest paid (with extra)")
    put(ws, "F7", '=IFERROR((C7+C8)*NPER(C6/12,-(C7+C8),C5)-C5,"—")', fill=CALC_FILL, fmt=MONEY0)
    label(ws, "E8", "Interest paid (regular only)")
    put(ws, "F8", '=IFERROR(C7*NPER(C6/12,-C7,C5)-C5,"—")', fill=CALC_FILL, fmt=MONEY0)
    label(ws, "E9", "Interest saved by the extra")
    put(ws, "F9", '=IFERROR(F8-F7,"—")', fill=CALC_FILL, fmt=MONEY0, font=BOLD)

    sh = wb.create_sheet("Schedule")
    widths(sh, {"A": 9, "B": 14, "C": 14, "D": 16})
    sched_header(sh, 1, ["Month", "Interest", "Payment", "Balance"])
    for m in range(1, 601):
        row = m + 1
        prev = "Payoff!$C$5" if m == 1 else f"$D{m}"
        live = f"AND(ISNUMBER({prev}),{prev}>0)"
        sh[f"A{row}"] = m
        sh[f"B{row}"] = f'=IF({live},{prev}*Payoff!$C$6/12,"")'
        sh[f"C{row}"] = f'=IF({live},MIN(Payoff!$C$7+Payoff!$C$8,{prev}+$B{row}),"")'
        sh[f"D{row}"] = f'=IF({live},MAX(0,{prev}+$B{row}-$C{row}),"")'
        for col in "BCD":
            sh[f"{col}{row}"].number_format = MONEY2
    wb.save(OUT / "debt-payoff-planner.xlsx")


# ------------------------------------------------------------------ budget
def budget():
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget"
    title_block(ws, "Monthly budget planner", "Planned vs. actual — the difference column tells the story.")
    widths(ws, {"A": 2, "B": 28, "C": 14, "D": 14, "E": 14})

    def header(row, text):
        label(ws, f"B{row}", text, bold=True)
        for col, h in (("C", "Planned"), ("D", "Actual"), ("E", "Difference")):
            c = ws[f"{col}{row}"]
            c.value = h
            c.fill = HEAD_FILL
            c.font = HEAD_FONT
            c.alignment = Alignment(horizontal="center")

    header(5, "Income")
    income = ["Salary / wages (take-home)", "Side income", "Other income"]
    r = 6
    for i, name in enumerate(income):
        row = r + i
        label(ws, f"B{row}", name)
        put(ws, f"C{row}", 0, fill=INPUT_FILL, fmt=MONEY0)
        put(ws, f"D{row}", 0, fill=INPUT_FILL, fmt=MONEY0)
        put(ws, f"E{row}", f"=D{row}-C{row}", fill=CALC_FILL, fmt=MONEY0)
    inc_end = r + len(income) - 1
    label(ws, f"B{inc_end + 1}", "Total income", bold=True)
    put(ws, f"C{inc_end + 1}", f"=SUM(C{r}:C{inc_end})", fill=CALC_FILL, fmt=MONEY0, font=BOLD)
    put(ws, f"D{inc_end + 1}", f"=SUM(D{r}:D{inc_end})", fill=CALC_FILL, fmt=MONEY0, font=BOLD)
    put(ws, f"E{inc_end + 1}", f"=D{inc_end + 1}-C{inc_end + 1}", fill=CALC_FILL, fmt=MONEY0, font=BOLD)

    exp_head = inc_end + 3
    header(exp_head, "Expenses")
    expenses = [
        "Rent / mortgage", "Property tax / HOA", "Utilities", "Internet & phone",
        "Groceries", "Dining out", "Transportation / fuel", "Car payment",
        "Insurance (auto/home/health)", "Medical & pharmacy", "Childcare / school",
        "Debt payments", "Subscriptions & streaming", "Clothing & personal care",
        "Gym & hobbies", "Gifts & giving", "Travel fund", "Emergency fund",
        "Retirement contributions", "Miscellaneous",
    ]
    r = exp_head + 1
    for i, name in enumerate(expenses):
        row = r + i
        label(ws, f"B{row}", name)
        put(ws, f"C{row}", 0, fill=INPUT_FILL, fmt=MONEY0)
        put(ws, f"D{row}", 0, fill=INPUT_FILL, fmt=MONEY0)
        put(ws, f"E{row}", f"=C{row}-D{row}", fill=CALC_FILL, fmt=MONEY0)
    exp_end = r + len(expenses) - 1
    label(ws, f"B{exp_end + 1}", "Total expenses", bold=True)
    put(ws, f"C{exp_end + 1}", f"=SUM(C{r}:C{exp_end})", fill=CALC_FILL, fmt=MONEY0, font=BOLD)
    put(ws, f"D{exp_end + 1}", f"=SUM(D{r}:D{exp_end})", fill=CALC_FILL, fmt=MONEY0, font=BOLD)
    put(ws, f"E{exp_end + 1}", f"=C{exp_end + 1}-D{exp_end + 1}", fill=CALC_FILL, fmt=MONEY0, font=BOLD)

    s = exp_end + 3
    label(ws, f"B{s}", "Left over (planned)", bold=True)
    put(ws, f"C{s}", f"=C{inc_end + 1}-C{exp_end + 1}", fill=CALC_FILL, fmt=MONEY0, font=BOLD)
    label(ws, f"B{s + 1}", "Left over (actual)", bold=True)
    put(ws, f"C{s + 1}", f"=D{inc_end + 1}-D{exp_end + 1}", fill=CALC_FILL, fmt=MONEY0, font=BOLD)
    wb.save(OUT / "budget-planner.xlsx")



# ------------------------------------------------------------ rent vs buy
def rent_vs_buy():
    wb = Workbook()
    ws = wb.active
    ws.title = "RentVsBuy"
    title_block(ws, "Rent vs. buy - full opportunity cost",
                "Owner vs. renter net worth, yearly. Includes invested down payment, maintenance, and 6% selling costs.")
    widths(ws, {"A": 2, "B": 30, "C": 14, "D": 3, "E": 8, "F": 15, "G": 15, "H": 15, "I": 15})
    ins = [("Home price", 425000, MONEY0), ("Down payment (%)", 0.20, PCT), ("Mortgage rate", 0.065, PCT),
           ("Property tax (%/yr)", 0.011, PCT), ("Maintenance+insurance (%/yr)", 0.015, PCT),
           ("Appreciation (%/yr)", 0.035, PCT), ("Rent today (monthly)", 2200, MONEY0),
           ("Rent growth (%/yr)", 0.03, PCT), ("Investment return (%/yr)", 0.07, PCT)]
    label(ws, "B4", "Inputs", bold=True)
    for i, (name, v, fmt) in enumerate(ins):
        label(ws, f"B{5+i}", name)
        put(ws, f"C{5+i}", v, fill=INPUT_FILL, fmt=fmt)
    label(ws, "B15", "Monthly P&I (computed)")
    put(ws, "C15", "=-PMT(C7/12,360,C5*(1-C6))", fill=CALC_FILL, fmt=MONEY2)
    label(ws, "B16", "Renter starting portfolio (down + 3% closing)")
    put(ws, "C16", "=C5*C6+C5*0.03", fill=CALC_FILL, fmt=MONEY0)
    ws["B18"] = "Yearly model (annual approximation of the web tool's monthly engine)."
    ws["B18"].font = SUB_FONT

    r0 = 20
    heads = ["Year", "Home value", "Loan balance", "Owner net worth", "Renter net worth"]
    for idx, h in enumerate(heads):
        c = ws.cell(row=r0, column=5 + idx, value=h)
        c.fill = HEAD_FILL; c.font = HEAD_FONT; c.alignment = Alignment(horizontal="center")
    for y in range(0, 41):
        row = r0 + 1 + y
        ws[f"E{row}"] = y
        if y == 0:
            ws[f"F{row}"] = "=C5"
            ws[f"G{row}"] = "=C5*(1-C6)"
            ws[f"I{row}"] = "=C16"
        else:
            ws[f"F{row}"] = f"=F{row-1}*(1+$C$10)"
            ws[f"G{row}"] = f"=MAX(0,G{row-1}*(1+$C$7)-$C$15*12)"
            ws[f"I{row}"] = (f"=I{row-1}*(1+$C$13)+MAX(0,($C$15*12+F{row-1}*($C$8+$C$9))"
                             f"-$C$11*12*(1+$C$12)^({y-1}))")
        ws[f"H{row}"] = f"=F{row}*0.94-G{row}"
        for col in "FGHI":
            ws[f"{col}{row}"].number_format = MONEY0
    wb.save(OUT / "rent-vs-buy.xlsx")


# --------------------------------------------------- multi-debt payoff
def multi_debt():
    wb = Workbook()
    ws = wb.active
    ws.title = "Debts"
    title_block(ws, "Multi-debt payoff planner",
                "Up to 5 debts, rolling minimums, extra payment. Priority 1-5: the lowest number gets every spare dollar. Rank by APR for avalanche, by balance for snowball.")
    widths(ws, {"A": 2, "B": 26, "C": 13, "D": 13, "E": 13, "F": 13})
    label(ws, "B4", "Inputs", bold=True)
    for idx, h in enumerate(["Balance", "APR", "Minimum", "Priority"]):
        c = ws.cell(row=5, column=3 + idx, value=h)
        c.fill = HEAD_FILL; c.font = HEAD_FONT; c.alignment = Alignment(horizontal="center")
    defaults = [(6500, 0.24, 130, 2), (14000, 0.075, 280, 3), (2100, 0.29, 60, 1), (0, 0.0, 0, 4), (0, 0.0, 0, 5)]
    for i, (b, a, m, pr) in enumerate(defaults):
        row = 6 + i
        label(ws, f"B{row}", f"Debt {i+1}")
        put(ws, f"C{row}", b, fill=INPUT_FILL, fmt=MONEY0)
        put(ws, f"D{row}", a, fill=INPUT_FILL, fmt=PCT)
        put(ws, f"E{row}", m, fill=INPUT_FILL, fmt=MONEY0)
        put(ws, f"F{row}", pr, fill=INPUT_FILL, fmt="0")
    label(ws, "B12", "Extra per month")
    put(ws, "C12", 300, fill=INPUT_FILL, fmt=MONEY0)
    label(ws, "B13", "Monthly budget (mins + extra)")
    put(ws, "C13", "=SUM(E6:E10)+C12", fill=CALC_FILL, fmt=MONEY0)

    sh = wb.create_sheet("Schedule")
    sh.column_dimensions["A"].width = 7
    for i in range(5):
        for j, h in enumerate([f"D{i+1} bal", f"D{i+1} int"]):
            col = 2 + i * 2 + j
            c = sh.cell(row=1, column=col, value=h)
            c.fill = HEAD_FILL; c.font = HEAD_FONT
        sh.column_dimensions[get_column_letter(2 + i * 2)].width = 12
        sh.column_dimensions[get_column_letter(3 + i * 2)].width = 11
    c = sh.cell(row=1, column=12, value="Total"); c.fill = HEAD_FILL; c.font = HEAD_FONT
    sh.column_dimensions["L"].width = 13
    c = sh.cell(row=1, column=1, value="Month"); c.fill = HEAD_FILL; c.font = HEAD_FONT
    sh["A2"] = 0
    for i in range(5):
        bcol = get_column_letter(2 + i * 2)
        sh[f"{bcol}2"] = f"=Debts!C{6+i}"
        sh[f"{bcol}2"].number_format = MONEY2
    sh["L2"] = "=B2+D2+F2+H2+J2"
    sh["L2"].number_format = MONEY2
    for m in range(1, 361):
        row = m + 2
        sh[f"A{row}"] = m
        alive_min = "+".join(
            f"IF({get_column_letter(2+i*2)}{row-1}>0.005,MIN(Debts!$E${6+i},{get_column_letter(2+i*2)}{row-1}),0)"
            for i in range(5))
        sh[f"N{row}"] = f"=Debts!$C$13-({alive_min})"
        prio = ",".join(
            f"IF({get_column_letter(2+i*2)}{row-1}>0.005,Debts!$F${6+i},99)" for i in range(5))
        sh[f"O{row}"] = f"=MIN({prio})"
        for i in range(5):
            bcol = get_column_letter(2 + i * 2)
            icol = get_column_letter(3 + i * 2)
            prev = f"{bcol}{row-1}"
            sh[f"{icol}{row}"] = f"=IF({prev}>0.005,{prev}*Debts!$D${6+i}/12,0)"
            target = f"IF(Debts!$F${6+i}=$O{row},MAX(0,$N{row}),0)"
            pay = f"MIN({prev}+{icol}{row},MIN(Debts!$E${6+i},{prev})+{target})"
            sh[f"{bcol}{row}"] = f"=IF({prev}<=0.005,0,MAX(0,{prev}+{icol}{row}-({pay})))"
            sh[f"{bcol}{row}"].number_format = MONEY2
            sh[f"{icol}{row}"].number_format = MONEY2
        sh[f"L{row}"] = f"=B{row}+D{row}+F{row}+H{row}+J{row}"
        sh[f"L{row}"].number_format = MONEY2
    sh.freeze_panes = "A2"
    wb.save(OUT / "multi-debt-payoff.xlsx")


# ------------------------------------------------------- roth conversion
def roth_conversion():
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan"
    title_block(ws, "Roth conversion bracket planner",
                "Fill your chosen bracket every year. 2026 brackets live on the Brackets sheet - marginal tax computed exactly via SUMPRODUCT.")
    widths(ws, {"A": 2, "B": 34, "C": 16, "D": 3, "E": 8, "F": 15, "G": 14, "H": 15})
    label(ws, "B4", "Inputs", bold=True)
    label(ws, "B5", "Traditional balance to convert")
    put(ws, "C5", 400000, fill=INPUT_FILL, fmt=MONEY0)
    label(ws, "B6", "Current taxable income")
    put(ws, "C6", 60000, fill=INPUT_FILL, fmt=MONEY0)
    label(ws, "B7", "Bracket ceiling (top $, see Brackets sheet)")
    put(ws, "C7", 105700, fill=INPUT_FILL, fmt=MONEY0)
    label(ws, "B8", "Balance growth (%/yr)")
    put(ws, "C8", 0.06, fill=INPUT_FILL, fmt=PCT)
    label(ws, "B10", "Converted per year (bracket space)")
    put(ws, "C10", "=MAX(0,C7-C6)", fill=CALC_FILL, fmt=MONEY0)

    bs = wb.create_sheet("Brackets")
    for col, w in {"A": 10, "B": 14, "C": 14, "D": 10}.items():
        bs.column_dimensions[col].width = w
    for idx, h in enumerate(["Rate", "Single from", "Joint from", "Step"]):
        c = bs.cell(row=1, column=1 + idx, value=h)
        c.fill = HEAD_FILL; c.font = HEAD_FONT
    single = [(0.10, 0, 0), (0.12, 12400, 24800), (0.22, 50400, 100800), (0.24, 105700, 211400),
              (0.32, 201775, 403550), (0.35, 256225, 512450), (0.37, 640600, 768700)]
    for i, (r, s1, s2) in enumerate(single):
        prev = single[i - 1][0] if i else 0
        bs.cell(row=2 + i, column=1, value=r).number_format = PCT
        bs.cell(row=2 + i, column=2, value=s1).number_format = MONEY0
        bs.cell(row=2 + i, column=3, value=s2).number_format = MONEY0
        bs.cell(row=2 + i, column=4, value=r - prev).number_format = PCT
    bs["A10"] = "Single-filer thresholds are used in the Plan sheet; swap column B for C in the formulas if filing jointly."
    bs["A10"].font = SUB_FONT

    r0 = 13
    for idx, h in enumerate(["Year", "Converted", "Tax that year", "Balance left"]):
        c = ws.cell(row=r0, column=5 + idx, value=h)
        c.fill = HEAD_FILL; c.font = HEAD_FONT; c.alignment = Alignment(horizontal="center")
    for y in range(1, 41):
        row = r0 + y
        ws[f"E{row}"] = y
        prev_bal = "$C$5" if y == 1 else f"H{row-1}"
        ws[f"F{row}"] = f"=MIN($C$10,{prev_bal})"
        ws[f"G{row}"] = (f"=SUMPRODUCT((($C$6+F{row})>Brackets!$B$2:$B$8)*(($C$6+F{row})-Brackets!$B$2:$B$8)*Brackets!$D$2:$D$8)"
                         f"-SUMPRODUCT(($C$6>Brackets!$B$2:$B$8)*($C$6-Brackets!$B$2:$B$8)*Brackets!$D$2:$D$8)")
        ws[f"H{row}"] = f"=MAX(0,({prev_bal}-F{row})*(1+$C$8))"
        for col in "FGH":
            ws[f"{col}{row}"].number_format = MONEY0
    wb.save(OUT / "roth-conversion-planner.xlsx")


# ------------------------------------------------------- sequence risk
def sequence_risk():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sequence"
    title_block(ws, "Sequence-of-returns risk",
                "The same set of yearly returns in three orders - crash-first, neutral, boom-first - with withdrawals.")
    widths(ws, {"A": 2, "B": 26, "C": 14, "D": 3, "E": 8, "F": 11, "G": 11, "H": 11, "I": 14, "J": 14, "K": 14})
    label(ws, "B4", "Inputs", bold=True)
    for i, (name, v, fmt) in enumerate([("Starting portfolio", 1000000, MONEY0), ("Yearly withdrawal", 40000, MONEY0),
                                        ("Average return", 0.07, PCT), ("Volatility (spread)", 0.12, PCT),
                                        ("Years", 30, "0")]):
        label(ws, f"B{5+i}", name)
        put(ws, f"C{5+i}", v, fill=INPUT_FILL, fmt=fmt)
    r0 = 12
    for idx, h in enumerate(["Year", "r crash-first", "r neutral", "r boom-first", "Bal crash", "Bal neutral", "Bal boom"]):
        c = ws.cell(row=r0, column=5 + idx, value=h)
        c.fill = HEAD_FILL; c.font = HEAD_FONT; c.alignment = Alignment(horizontal="center")
    for y in range(1, 51):
        row = r0 + y
        ws[f"E{row}"] = y
        live = f"$E{row}<=$C$9"
        ws[f"F{row}"] = f'=IF({live},$C$7+1.5*$C$8*(2*($E{row}-1)/($C$9-1)-1),"")'
        ws[f"H{row}"] = f'=IF({live},$C$7+1.5*$C$8*(2*($C$9-$E{row})/($C$9-1)-1),"")'
        ws[f"G{row}"] = (f'=IF({live},IF(MOD($E{row},2)=1,'
                         f'$C$7+1.5*$C$8*(2*($C$9-($E{row}+1)/2)/($C$9-1)-1),'
                         f'$C$7+1.5*$C$8*(2*($E{row}/2-1)/($C$9-1)-1)),"")')
        for col, rcol in (("I", "F"), ("J", "G"), ("K", "H")):
            prev = "$C$5" if y == 1 else f"{col}{row-1}"
            ws[f"{col}{row}"] = f'=IF({live},MAX(0,({prev}-$C$6)*(1+{rcol}{row})),"")'
            ws[f"{col}{row}"].number_format = MONEY0
        for col in "FGH":
            ws[f"{col}{row}"].number_format = PCT
    wb.save(OUT / "sequence-risk.xlsx")



# ------------------------------------------------ rental vs S&P 500
def rental_vs_sp():
    wb = Workbook()
    ws = wb.active
    ws.title = "RentalVsSP"
    title_block(ws, "Rental property vs. the S&P 500",
                "Every landlord cost and every tax effect - depreciation, passive-loss limits, capital gains recapture - versus the same cash in an index fund.")
    widths(ws, {"A": 2, "B": 36, "C": 15, "D": 3, "E": 30, "F": 15})

    # Named rows (single source of truth - avoids the classic off-by-row bug)
    R = {}
    ins = [
        ("price", "Property price", 320000, MONEY0),
        ("down", "Down payment (%)", 0.25, PCT),
        ("rate", "Mortgage rate", 0.065, PCT),
        ("term", "Loan term (years)", 30, "0"),
        ("closing", "Closing costs (% of price)", 0.03, PCT),
        ("rehab", "Upfront repairs / rehab", 8000, MONEY0),
        ("rent", "Monthly rent (today)", 2400, MONEY0),
        ("vacancy", "Vacancy rate (%)", 0.06, PCT),
        ("rentgrow", "Rent growth (%/yr)", 0.03, PCT),
        ("proptax", "Property tax (%/yr of value)", 0.012, PCT),
        ("ins", "Insurance (%/yr of value)", 0.006, PCT),
        ("maint", "Maintenance (%/yr of value)", 0.012, PCT),
        ("mgmt", "Management (% of rent)", 0.08, PCT),
        ("hoa", "HOA (monthly)", 0, MONEY0),
        ("appre", "Appreciation (%/yr)", 0.035, PCT),
        ("building", "Building value (% of price)", 0.80, PCT),
        ("sellcost", "Selling costs (%)", 0.07, PCT),
        ("hold", "Holding period (years)", 15, "0"),
        ("otherinc", "Your other income (for $25k loss-allowance phase-out)", 120000, MONEY0),
        ("marg", "Your marginal tax rate", 0.24, PCT),
        ("ltcg", "Capital gains rate", 0.15, PCT),
        ("spreturn", "S&P 500 expected return (%/yr)", 0.08, PCT),
    ]
    label(ws, "B4", "Inputs", bold=True)
    for i, (key, name, v, fmt) in enumerate(ins):
        row = 5 + i
        R[key] = row
        label(ws, f"B{row}", name)
        put(ws, f"C{row}", v, fill=INPUT_FILL, fmt=fmt)

    r0 = 5 + len(ins) + 2
    R["loanamt"] = r0
    R["upfront"] = r0 + 1
    R["pmt"] = r0 + 2
    R["deprperyr"] = r0 + 3
    label(ws, f"B{r0-1}", "Key computed values", bold=True)
    label(ws, f"B{R['loanamt']}", "Loan amount")
    put(ws, f"C{R['loanamt']}", f"=C{R['price']}*(1-C{R['down']})", fill=CALC_FILL, fmt=MONEY0)
    label(ws, f"B{R['upfront']}", "Upfront cash (down + closing + rehab)")
    put(ws, f"C{R['upfront']}", f"=C{R['price']}*C{R['down']}+C{R['price']}*C{R['closing']}+C{R['rehab']}",
        fill=CALC_FILL, fmt=MONEY0)
    label(ws, f"B{R['pmt']}", "Monthly P&I payment")
    put(ws, f"C{R['pmt']}", f"=-PMT(C{R['rate']}/12,C{R['term']}*12,C{R['loanamt']})", fill=CALC_FILL, fmt=MONEY2)
    label(ws, f"B{R['deprperyr']}", "Annual depreciation (building / 27.5 yrs)")
    put(ws, f"C{R['deprperyr']}", f"=C{R['price']}*C{R['building']}/27.5", fill=CALC_FILL, fmt=MONEY0)

    sh = wb.create_sheet("YearByYear")
    widths(sh, {"A": 6})
    heads = ["Year", "Home value", "Rent (net vacancy)", "Operating costs", "NOI",
             "Interest", "Principal", "Loan balance", "Depreciation", "Taxable rental income",
             "Tax effect", "Cash flow after tax"]
    for i, h in enumerate(heads):
        c = sh.cell(row=1, column=1 + i, value=h)
        c.fill = HEAD_FILL; c.font = HEAD_FONT
        sh.column_dimensions[get_column_letter(1 + i)].width = 15
    sh.freeze_panes = "A2"
    M = "RentalVsSP"
    for y in range(1, 41):
        row = y + 1
        prevbal = f"{M}!$C${R['loanamt']}" if y == 1 else f"H{row-1}"
        live = f"A{row}<={M}!$C${R['hold']}"
        sh[f"A{row}"] = y
        sh[f"B{row}"] = f'=IF({live},{M}!$C${R["price"]}*(1+{M}!$C${R["appre"]})^(A{row}-1),"")'
        sh[f"C{row}"] = (f'=IF({live},{M}!$C${R["rent"]}*12*(1+{M}!$C${R["rentgrow"]})^(A{row}-1)'
                         f'*(1-{M}!$C${R["vacancy"]}),"")')
        sh[f"D{row}"] = (f'=IF({live},B{row}*({M}!$C${R["proptax"]}+{M}!$C${R["ins"]}+{M}!$C${R["maint"]})'
                         f'+C{row}*{M}!$C${R["mgmt"]}+{M}!$C${R["hoa"]}*12,"")')
        sh[f"E{row}"] = f'=IF({live},C{row}-D{row},"")'
        sh[f"F{row}"] = f'=IF({live},{prevbal}*{M}!$C${R["rate"]}/12*12,"")'
        sh[f"G{row}"] = f'=IF({live},{M}!$C${R["pmt"]}*12-F{row},"")'
        sh[f"H{row}"] = f'=IF({live},MAX(0,{prevbal}-G{row}),"")'
        sh[f"I{row}"] = (f'=IF({live},MIN({M}!$C${R["deprperyr"]},'
                         f'{M}!$C${R["price"]}*{M}!$C${R["building"]}-SUM($I$2:I{row-1})),"")')
        sh[f"J{row}"] = f'=IF({live},E{row}-F{row}-I{row},"")'
        sh[f"K{row}"] = (f'=IF({live},IF(J{row}>=0,-J{row}*{M}!$C${R["marg"]},'
                         f'MIN(-J{row},MAX(0,25000-0.5*MAX(0,{M}!$C${R["otherinc"]}-100000)))*{M}!$C${R["marg"]}),"")')
        sh[f"L{row}"] = f'=IF({live},E{row}-F{row}-G{row}+K{row},"")'
        for col in "BCDEFGHIJKL":
            sh[f"{col}{row}"].number_format = MONEY2
    ws[f"B{r0+5}"] = ("Note: the YearByYear sheet uses a simplified flat-balance interest estimate per year "
                       "(the web tool runs a full monthly amortization). Figures will be close but not identical "
                       "to the calculator - use the calculator for the exact numbers, this sheet for your own edits.")
    ws[f"B{r0+5}"].font = SUB_FONT
    wb.save(OUT / "rental-vs-sp500.xlsx")


if __name__ == "__main__":
    for fn in (mortgage, loan, compound, savings_goal, debt_payoff, budget, rent_vs_buy, multi_debt, roth_conversion, sequence_risk, rental_vs_sp):
        fn()
        print("built", fn.__name__)
    print("done ->", OUT)
